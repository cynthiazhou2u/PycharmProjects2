# Python program to read an excel file
print('\f1st menthod\f')
# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Getting the value of maximum rows
# and column
row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# printing the value of first column
# Loop will print all values
# of first column
print("\nValue of first column")
for i in range(1, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value)

# printing the value of second row
# Loop will print all values
# of first row
print("\nValue of second row")
for i in range(1, column + 1):
    cell_obj = sheet_obj.cell(row=2, column=i)
    print(cell_obj.value, end=" ")


print('\n\n\f\f\f 2nd method\f\f\f')
# Python program to read an excel file

# import openpyxl module
import openpyxl

# Give the location of the file
path = "gfg.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell object is created by using
# sheet object's cell() method.
cell_obj = sheet_obj['A1': 'D6']

# Print value of cell object
# using the value attribute
for cell1, cell2, cell3, cell4 in cell_obj:
	print(cell1.value, cell2.value, cell3.value, cell4.value)

print('\n\f\f\f3rd method\f\f\f')
# Iterate the loop to read the cell values
import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook("gfg.xlsx")

# Define variable to read sheet
dataframe1 = dataframe.active

for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(col[row].value,end=' ')
    print('\n')

print('\n\f\f\f4th method\f\f\f')

# Python3 code to select
# data from excel
import xlwings as xw

# Specifying a sheet
ws = xw.Book("gfg.xlsx").sheets['Sheet1']

# Selecting data from
# a single cell
v1 = ws.range("A1:D6").value
#v2 = ws.range("A1:D6").value
print("Result:\n", v1)
print('OR===')
for i in range(0,6):
    print(v1[i])

print('\n\f\f\f5th method\f\f\f')

# import pandas lib as pd
import pandas as pd

# read by default 1st sheet of an excel file
dataframe1 = pd.read_excel('gfg.xlsx')

print(dataframe1)